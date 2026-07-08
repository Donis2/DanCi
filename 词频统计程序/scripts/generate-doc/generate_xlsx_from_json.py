# -*- coding: utf-8 -*-
"""
从 netem_full_list.json 生成 Excel 文档 (不依赖 MySQL)

特性:
  - 表头加粗 + 灰底 + 居中
  - 首行冻结
  - 自动列宽
  - 自动筛选
  - 词频列右对齐

运行: python scripts/generate-doc/generate_xlsx_from_json.py
"""

import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
INPUT_JSON = os.path.join(BASE, "netem_full_list.json")
OUTPUT_XLSX = os.path.join(BASE, "netem_full_list.xlsx")

# 列定义: (列名, json_key, 列宽)
COLUMNS = [
    ("序号", "序号", 6),
    ("词频", "词频", 8),
    ("单词", "单词", 16),
    ("释义", "释义", 50),
    ("其他拼写", "其他拼写", 16),
    ("分类", "分类", 12),
    ("子分类", "子分类", 14),
]


def main():
    # 1. 读取 JSON
    print(f"[1/3] 读取 {INPUT_JSON}")
    with open(INPUT_JSON, 'r', encoding='utf-8') as f:
        data = json.load(f)
    key = list(data.keys())[0]
    rows = data[key]
    total = len(rows)
    print(f"      共 {total} 条记录,标题: {key}")

    # 2. 生成 Excel
    print("[2/3] 生成 Excel")
    wb = Workbook()
    ws = wb.active
    ws.title = key[:31]  # Excel sheet 名最多 31 字符

    # 样式定义
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name="微软雅黑", size=10)
    body_align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    body_align_right = Alignment(horizontal="right", vertical="center")
    body_align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # 写表头
    for ci, (col_name, _, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 写数据
    for ri, row_data in enumerate(rows, 2):
        for ci, (_, json_key, _) in enumerate(COLUMNS, 1):
            value = row_data.get(json_key)
            if value is None:
                value = ""
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font = body_font
            cell.border = thin_border
            # 对齐: 序号/词频右对齐, 单词居中, 其余左对齐
            if json_key in ("序号", "词频"):
                cell.alignment = body_align_right
            elif json_key == "单词":
                cell.alignment = body_align_center
            else:
                cell.alignment = body_align_left

    # 列宽
    for ci, (_, _, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{total + 1}"

    # 3. 保存
    print(f"[3/3] 保存 {OUTPUT_XLSX}")
    wb.save(OUTPUT_XLSX)
    print(f"完成,共 {total} 行数据")


if __name__ == "__main__":
    main()
